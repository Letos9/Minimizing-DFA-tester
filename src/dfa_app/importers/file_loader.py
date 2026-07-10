from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from dfa_app.domain.models import DFA
from dfa_app.importers.dot_parser import DotParseError, parse_dot
from dfa_app.importers.parser import EXPECTED_HEADERS, parse_row


@dataclass(frozen=True, slots=True)
class RowError:
    row_number: int
    message: str


@dataclass(frozen=True, slots=True)
class LoadedAutomata:
    automata: tuple[DFA, ...]
    errors: tuple[RowError, ...]


def _normalize_headers(values: Iterable[object]) -> tuple[str, ...]:
    return tuple("" if value is None else str(value).strip().lower() for value in values)


def _parse_rows(headers: tuple[str, ...], rows: Iterable[tuple[int, Iterable[object]]]) -> LoadedAutomata:
    if headers != EXPECTED_HEADERS:
        expected = "; ".join(EXPECTED_HEADERS)
        actual = "; ".join(headers) or "<пусто>"
        return LoadedAutomata((), (RowError(1, f"неверные заголовки: {actual}; ожидается: {expected}"),))

    automata: list[DFA] = []
    errors: list[RowError] = []
    for row_number, values in rows:
        values = tuple(values)
        if not any(value is not None and str(value).strip() for value in values):
            continue
        row = dict(zip(headers, values, strict=False))
        try:
            automata.append(parse_row(row))
        except ValueError as exc:
            errors.append(RowError(row_number, str(exc)))
    return LoadedAutomata(tuple(automata), tuple(errors))


def _load_delimited(path: Path) -> LoadedAutomata:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream, delimiter=";")
            try:
                headers = _normalize_headers(next(reader))
            except StopIteration:
                return LoadedAutomata((), (RowError(1, "файл пуст"),))
            return _parse_rows(headers, ((number, row) for number, row in enumerate(reader, start=2)))
    except UnicodeDecodeError:
        return LoadedAutomata((), (RowError(1, "файл должен иметь кодировку UTF-8"),))


def _load_excel(path: Path) -> LoadedAutomata:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return LoadedAutomata((), (RowError(1, f"не удалось прочитать Excel: {exc}"),))
    try:
        rows = workbook.active.iter_rows(values_only=True)
        try:
            headers = _normalize_headers(next(rows))
        except StopIteration:
            return LoadedAutomata((), (RowError(1, "файл пуст"),))
        return _parse_rows(headers, ((number, row) for number, row in enumerate(rows, start=2)))
    finally:
        workbook.close()


def _load_dot(path: Path) -> LoadedAutomata:
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        return LoadedAutomata((), (RowError(1, "файл должен иметь кодировку UTF-8"),))
    try:
        return LoadedAutomata((parse_dot(text),), ())
    except DotParseError as exc:
        return LoadedAutomata((), (RowError(exc.line_number, str(exc)),))


def load_automata(file_path: str | Path) -> LoadedAutomata:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in {".txt", ".csv"}:
        return _load_delimited(path)
    if suffix == ".xlsx":
        return _load_excel(path)
    if suffix in {".dot", ".gv"}:
        return _load_dot(path)
    return LoadedAutomata((), (RowError(1, f"неподдерживаемый формат файла: {suffix or '<без расширения>'}"),))
